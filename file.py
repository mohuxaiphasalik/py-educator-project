from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill,Color,Font,colors
from calendar import monthrange
wb = openpyxl.load_workbook("1.xlsx")
ft = Font(color="F50707")
s1 = wb.active
max_col = s1.max_column
max_row = s1.max_row
########################################################################
#functions
def remove_double():
    for i in range(2, max_row + 1):
        if s1.cell(i, 5).value == s1.cell(i + 1, 5).value:
            s1.delete_rows(i, 1)
        elif s1.cell(i, 5).value == s1.cell(i - 1, 5).value:
            s1.delete_rows(i, 1)
    wb.save("2.xlsx")  
# deleting empty rows
def delete_empty():
    for i in range(2,max_row+1):
        if s1.cell(i,5).value==None:
            s1.delete_rows(i,1)
    wb.save("2.xlsx")
#finding absents
def find_absents(i,day, total_days_in_m):
    j = 1
    d=start_date
    name=s1.cell(i, 2).value
    first_date=day
    while day<=total_days_in_m:
        if s1.cell(i, 5).value == "In" or s1.cell(i, 5).value == "Break":
            if int(s1.cell(i, 3).value.strftime("%d")) != day:
                s1.insert_rows(i, 1)
                s1.cell(i,2).value=f"Absent---> {day}/{month}/{year}"
                s1.cell(i,2).fill=PatternFill("solid",fgColor="F50707") 
            day += 1            
            #d=d+timedelta(days=1)
        i+=1
    if s1.cell(i,2).value != name:
        s1.insert_rows(i)
        s1.cell(i, 2).fill = PatternFill("solid", fgColor="71FF33")
        wb.save("2.xlsx")
    else:
        i=i+1
    if s1.cell(i+1, 2).value != None :
        i=i+1
        find_absents(i,first_date,total_days_in_m)
    else:
        return    
    wb.save("2.xlsx")


def late_check(ontime_time):
    for i in range(2,max_row_1+1):
        if s2.cell(i, 3).value!=None:
            if s2.cell(i, 5).value == "In" or s2.cell(i, 5).value == "Break":
                time=s2.cell(i, 3).value

                if datetime.strptime(("%s:%s" % (time.hour,time.minute)),"%H:%M") > ontime_time:
                    s2.cell(i,3).font=ft
    
    wb1.save("2.xlsx")



###################################################    ```````````main`      
ontime_input = input("Enter Ontime (format Hour:Minute) :")
start_date = input("Enter start Date (format=dd/mm/yyyy): ")
ontime_time=datetime.strptime(ontime_input, "%H:%M")
complete_time_str = f"{start_date} {ontime_input}"
complete_f = datetime.strptime(complete_time_str, "%d/%m/%Y %H:%M")
day = int(complete_f.strftime("%d"))
year = int(complete_f.strftime("%Y"))
month = int(complete_f.strftime("%m"))
total_days_in_m = monthrange(year, month)[1]


wb1 = openpyxl.load_workbook("2.xlsx")

s2 = wb1.active
max_col_1 = s2.max_column
max_row_1 = s2.max_row


remove_double()
remove_double()
delete_empty()
#find_absents(2,day, total_days_in_m)
late_check(ontime_time)
print("done")
